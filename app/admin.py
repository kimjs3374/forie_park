"""관리사무소(admin)용 화면 — 가입 승인 / 반려 및 전체 방문등록 조회."""
import csv
import io          # CSV 내보내기에서 쓴다
import re
from urllib.parse import quote

from markupsafe import Markup
from datetime import datetime, timezone, timedelta
from functools import wraps

from flask import (Blueprint, render_template, redirect, url_for, flash, request,
                   abort, Response, current_app)
from flask_login import login_required, current_user

from . import models
from . import analytics
from . import usage

admin_bp = Blueprint("admin", __name__, url_prefix="/admin")


def admin_required(f):
    @wraps(f)
    @login_required
    def wrapper(*args, **kwargs):
        if not current_user.is_admin:
            abort(403)
        return f(*args, **kwargs)
    return wrapper


@admin_bp.route("/")
@admin_required
def dashboard():
    # 계정·명부는 통합 관리(forie.kr/admin)의 몫이다. 여기서는 방문차량만 센다.
    visit_count = models.visits_count(status="active")
    try:
        suspect_count = analytics.scan_cached()["stats"]["flagged"]
    except Exception:
        # 분석이 실패해도 대시보드 자체는 열려야 한다.
        current_app.logger.exception("의심세대 집계 실패")
        suspect_count = 0
    try:
        overuse_count = usage.scan_overuse_cached()["stats"]["flagged"]
    except Exception:
        current_app.logger.exception("실주차일수 초과 집계 실패")
        overuse_count = 0
    return render_template("admin/dashboard.html",
                           visit_count=visit_count, suspect_count=suspect_count,
                           overuse_count=overuse_count)


# 계정·명부 관리는 통합 관리(forie.kr/admin)로 이관되었다.
# 옛 주소를 여기 남겨두면 '이 앱에서도 되는 기능'처럼 보여 오히려 헷갈린다.


def _parse_date(value):
    try:
        return datetime.strptime(value, "%Y-%m-%d")
    except (ValueError, TypeError):
        return None


# 한 번에 그리는 방문등록 카드 수. 카드마다 로그 타임라인이 붙어 무겁기에
# 첫 화면을 가볍게 하고 나머지는 스크롤에 맞춰 이어 붙인다.
PAGE_SIZE = 40


def _visit_view(regs):
    """카드 조각이 필요로 하는 부속 자료(로그 요약·표시용 입출차)."""
    logs_map = models.visit_logs_by_regs([r.id for r in regs])
    now_utc = datetime.now(timezone.utc)
    return {
        "summaries": {r.id: models.summarize_logs(logs_map.get(r.id, []), now=now_utc)
                      for r in regs},
        "dispmap": models.visit_dispmap(regs, now=now_utc, logs_map=logs_map),
    }


@admin_bp.route("/visits")
@admin_required
def visits():
    date_from = _parse_date(request.args.get("from"))
    date_to = _parse_date(request.args.get("to"))
    car = request.args.get("car", "")
    regs = models.visits_filter(date_from, date_to, limit=PAGE_SIZE, with_user=True, car=car)
    return render_template(
        "admin/visits.html",
        registrations=regs,
        total=models.visits_count_filter(date_from, date_to, car),
        date_from=request.args.get("from", ""),
        date_to=request.args.get("to", ""),
        car=car,
        car_norm=models.normalize_car_query(car),
        **_visit_view(regs),
    )


@admin_bp.route("/visits/page")
@admin_required
def visits_page():
    """무한 스크롤이 이어 붙일 카드 조각. 남은 게 없으면 빈 응답."""
    date_from = _parse_date(request.args.get("from"))
    date_to = _parse_date(request.args.get("to"))
    car = request.args.get("car", "")
    try:
        offset = max(0, int(request.args.get("offset", 0)))
    except (TypeError, ValueError):
        offset = 0
    regs = models.visits_filter(date_from, date_to, limit=PAGE_SIZE, with_user=True,
                                car=car, offset=offset)
    if not regs:
        return ""
    return render_template("admin/_visit_items.html", registrations=regs, **_visit_view(regs))


# 등록 상태/연동 상태 한글 라벨
_STATUS_KO = {"active": "유효", "cancelled": "취소"}
_SYNC_KO = {"pending": "전송대기", "synced": "연동완료", "failed": "실패"}


_STATE_KO = {"entered": "주차중", "exited": "출차완료"}
_EVENT_KO = {"in": "입차", "out": "출차"}


def _kst(dt):
    """timestamptz(UTC) → KST naive. 실제 입출차·로그 이벤트용."""
    return analytics._kst(dt)


def _fmt(dt, pattern="%Y-%m-%d %H:%M"):
    return dt.strftime(pattern) if dt else ""


def _csv_response(rows, header, prefix):
    buf = io.StringIO()
    buf.write("﻿")  # Excel 한글 깨짐 방지용 UTF-8 BOM
    writer = csv.writer(buf)
    writer.writerow(header)
    writer.writerows(rows)
    today = datetime.now().strftime("%Y%m%d")
    return Response(
        buf.getvalue(),
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={prefix}_{today}.csv"},
    )


def _visit_rows(regs):
    """방문등록 → 표/CSV 공용 행. 실제 입출차는 KST 로 환산해 담는다."""
    out = []
    for r in regs:
        ain, aout = _kst(r.actual_in_time), _kst(r.actual_out_time)
        stay = round((aout - ain).total_seconds() / 60) if (ain and aout) else ""
        out.append([
            r.id, _fmt(r.created_at and _kst(r.created_at)),
            r.dong, r.ho, r.user_name or "", r.car_number, r.car_type or "",
            r._row.get("visit_reason") or "", r._row.get("visitor_phone") or "",
            _fmt(r.entry_time, "%Y-%m-%d"), _fmt(r.exit_time, "%Y-%m-%d"),
            _fmt(ain), _fmt(aout), stay,
            _STATUS_KO.get(r.status, r.status),
            _STATE_KO.get(r.visit_state, ""),
            "예" if r.nexpa_registered else "아니오",
            _SYNC_KO.get(r.nexpa_sync_status, r.nexpa_sync_status),
        ])
    return out


_VISIT_HEADER = [
    "등록ID", "등록일시", "동", "호", "신청자", "차량번호", "차량종류",
    "방문사유", "방문자연락처", "신청 입차일", "신청 출차일",
    "실제 입차(KST)", "실제 출차(KST)", "체류(분)",
    "등록상태", "방문상태", "넥스파등록", "연동상태",
]
_LOG_HEADER = [
    "로그ID", "등록ID", "동", "호", "신청자", "차량번호", "구분",
    "발생시각(KST)", "출처", "신청 입차일", "등록상태",
]


def _log_rows(logs, regs):
    """입출차 로그 → 행. 로그에는 세대가 없어 등록건을 붙여 채운다."""
    regmap = {r.id: r for r in regs}
    out = []
    for lg in logs:
        r = regmap.get(lg.registration_id)
        out.append([
            lg.id, lg.registration_id,
            r.dong if r else "", r.ho if r else "", (r.user_name or "") if r else "",
            lg.car_number, _EVENT_KO.get(lg.event_type, lg.event_type or ""),
            _fmt(lg.event_time_kst), lg.source,
            _fmt(r.entry_time, "%Y-%m-%d") if r else "",
            _STATUS_KO.get(r.status, r.status) if r else "",
        ])
    return out


@admin_bp.route("/visits/export")
@admin_required
def export_visits():
    """방문등록 내역 CSV. 기간·차량번호를 비우면 전체."""
    date_from = _parse_date(request.args.get("from"))
    date_to = _parse_date(request.args.get("to"))
    car = request.args.get("car", "")
    regs = models.visits_filter_all(date_from, date_to, with_user=True, car=car)
    return _csv_response(_visit_rows(regs), _VISIT_HEADER, "visit_registrations")


@admin_bp.route("/logs/export")
@admin_required
def export_logs():
    """입출차 로그 CSV — 기간 내 이벤트 전량(1000행 상한 없음)."""
    date_from = _parse_date(request.args.get("from"))
    date_to = _parse_date(request.args.get("to"))
    car = request.args.get("car", "")
    logs = models.visit_logs_filter(date_from, date_to, car=car)
    # 로그의 세대·신청자를 채우려면 등록건이 필요하다. 기간 밖 등록에 달린
    # 로그도 있을 수 있어 등록은 기간 제한 없이 전량 가져온다(차량 필터는 유지).
    regs = models.visits_filter_all(with_user=True, car=car)
    return _csv_response(_log_rows(logs, regs), _LOG_HEADER, "visit_logs")


@admin_bp.route("/export.xlsx")
@admin_required
def export_workbook():
    """방문등록·입출차로그·요약·의심세대를 한 파일에 담은 엑셀."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    date_from = _parse_date(request.args.get("from"))
    date_to = _parse_date(request.args.get("to"))
    car = request.args.get("car", "")
    regs = models.visits_filter_all(date_from, date_to, with_user=True, car=car)
    logs = models.visit_logs_filter(date_from, date_to, car=car)
    all_regs = models.visits_filter_all(with_user=True)
    # 의심세대는 세대 단위 분석이라 차량 검색어를 걸면 뜻이 흐려진다. 기간만 적용한다.
    report = analytics.scan(date_from, date_to)

    wb = Workbook()
    fill = PatternFill("solid", fgColor="1F3864")
    head = Font(color="FFFFFF", bold=True, size=10)

    def add(title, header, rows, widths):
        ws = wb.create_sheet(title)
        ws.append(header)
        for c in range(1, len(header) + 1):
            ws.cell(1, c).fill = fill
            ws.cell(1, c).font = head
            ws.cell(1, c).alignment = Alignment(horizontal="center", wrap_text=True)
        for row in rows:
            ws.append(row)
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{max(ws.max_row, 1)}"
        return ws

    add("방문등록", _VISIT_HEADER, _visit_rows(regs),
        [8, 17, 6, 7, 11, 12, 10, 16, 14, 12, 12, 17, 17, 9, 9, 9, 10, 9])
    add("입출차로그", _LOG_HEADER, _log_rows(logs, all_regs),
        [8, 8, 6, 7, 11, 12, 7, 18, 9, 12, 9])
    add("의심세대",
        ["순위", "세대", "계정", "점수", "위험도", "등록건수", "차량종수", "이용일수",
         "최대반복", "연속일", "야간", "누적체류(h)", "미입차", "주요차량", "판단근거"],
        [[i, h["household"], h["users"], h["score"], h["level"], h["regs"], h["cars"],
          h["days"], h["repeat"], h["run"], h["night"], h["stay_hours"], h["no_in"],
          h["top_car"], " / ".join(h["why"])]
         for i, h in enumerate(report["households"], 1)],
        [6, 12, 14, 7, 8, 9, 9, 9, 9, 8, 7, 11, 8, 12, 60])
    add("데이터이상", ["유형", "차량번호", "세대", "신청자", "등록ID", "내용"],
        [[a["kind"], a["car_number"], a["household"], a["user"], a["reg_id"], a["detail"]]
         for a in report["anomalies"]],
        [18, 12, 12, 10, 8, 46])

    ws = wb["Sheet"]
    ws.title = "개요"
    st = report["stats"]
    now_kst = _kst(datetime.now(timezone.utc))
    period = "%s ~ %s" % (request.args.get("from") or "처음",
                          request.args.get("to") or "오늘")
    for k, v in [
        ("항목", "값"),
        ("추출일시(KST)", _fmt(now_kst)),
        ("조회 기간", period),
        ("차량번호 검색", (models.normalize_car_query(car) + " (부분 일치)") if car else "전체"),
        # 방문등록·입출차로그 시트는 차량 검색까지 반영된 실제 행수를 적는다.
        ("방문등록", "%d건 (유효 %d건)" % (len(regs), sum(1 for r in regs if r.status != "cancelled"))),
        ("입출차 로그", "%d건" % len(logs)),
        ("이용 세대수", "%d세대" % st["households"]),
        ("의심 세대", "%d세대 (높음 %d / 중간 %d) — 기간 전체 기준"
                       % (st["flagged"], st["high"], st["mid"])),
        ("데이터 이상", "%d건" % len(report["anomalies"])),
        ("", ""),
        ("점수 산출식", report["scoring"]["formula"]),
        ("배점 근거", " · ".join("%s %s(%s, %s)" % (i["name"], i["weight"], i["unit"], i["free"])
                                  for i in report["scoring"]["items"])),
        ("위험도 구간", " · ".join("%s %s" % (lv["label"], lv["range"])
                                    for lv in report["scoring"]["levels"])),
        ("목록 하한", report["scoring"]["hidden_note"]),
        ("", ""),
        ("시간대 표기", "실제 입/출차 및 로그는 KST(UTC+9). 신청 입/출차일은 저장값 그대로(일 단위)."),
        ("주의", "연락처·차량번호가 포함된 개인정보 문서입니다. 취급에 유의하세요."),
        ("의심세대 점수", "살펴볼 순서를 정하는 참고값이며 위반 판정이 아닙니다. "
                          "정당한 사정이 있을 수 있으니 소명을 먼저 받으세요."),
    ]:
        ws.append([k, v])
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 78
    ws["A1"].fill = fill
    ws["A1"].font = head
    ws["B1"].fill = fill
    ws["B1"].font = head
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 1).font = Font(bold=True, size=10)
        ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="center")
    wb._sheets = [wb["개요"], wb["방문등록"], wb["입출차로그"], wb["의심세대"], wb["데이터이상"]]

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    today = datetime.now().strftime("%Y%m%d")
    return Response(
        out.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=parking_report_{today}.xlsx"},
    )


@admin_bp.route("/suspects")
@admin_required
def suspects():
    """부정사용(상시주차 우회) 의심 세대 목록."""
    date_from = _parse_date(request.args.get("from"))
    date_to = _parse_date(request.args.get("to"))
    # 기간을 안 좁힌 첫 화면은 대시보드 배지와 같은 집계다 — 캐시를 함께 쓴다.
    report = (analytics.scan_cached() if not (date_from or date_to)
              else analytics.scan(date_from, date_to))
    return render_template(
        "admin/suspects.html",
        households=report["households"],
        anomalies=report["anomalies"],
        stats=report["stats"],
        scoring=report["scoring"],
        date_from=request.args.get("from", ""),
        date_to=request.args.get("to", ""),
    )


@admin_bp.route("/suspects/export")
@admin_required
def export_suspects():
    date_from = _parse_date(request.args.get("from"))
    date_to = _parse_date(request.args.get("to"))
    report = analytics.scan(date_from, date_to)
    rows = [[i, h["household"], h["users"], h["score"], h["level"], h["regs"], h["cars"],
             h["days"], h["repeat"], h["run"], h["night"], h["stay_hours"], h["no_in"],
             h["top_car"], h["car_list"], h["period"], " / ".join(h["why"])]
            for i, h in enumerate(report["households"], 1)]
    header = ["순위", "세대", "계정", "점수", "위험도", "등록건수", "차량종수", "이용일수",
              "최대반복", "연속일", "야간체류", "누적체류(시간)", "미입차",
              "주요차량", "전체차량", "이용기간", "판단근거"]
    return _csv_response(rows, header, "parking_suspects")


@admin_bp.route("/overuse")
@admin_required
def overuse():
    """실주차일수(관제 입출차 기준) 월 한도 초과 차량."""
    report = usage.scan_overuse_cached(month=request.args.get("month"))
    # 달 선택기 — 이번 달부터 12개월 뒤로.
    base = usage.today_kst().replace(day=1)
    months = []
    for _ in range(12):
        months.append("%d-%02d" % (base.year, base.month))
        base = (base - timedelta(days=1)).replace(day=1)
    return render_template("admin/overuse.html", months=months, **report)


_OVERUSE_HEADER = ["순위", "차량번호", "숙박일수", "초과일수", "야간 누적주차",
                   "첫 숙박일", "마지막 숙박일", "최근 이벤트(KST)", "등록 세대",
                   "시스템등록", "미출차"]

_OPEN_HEADER = ["차량번호", "입차(KST)", "경과시간", "경과일", "현재 집계 숙박일수",
                "세대", "정기등록", "구분"]


@admin_bp.route("/overuse/open-export")
@admin_required
def export_open_stays():
    """미출차 점검 목록 CSV — 관제에 출차 누락을 확인시킬 때 쓴다."""
    rows = [[r["car_number"], _fmt(r["entered_at"]), r["hours"], r["days"],
             r["counted_nights"], r["household"],
             "예" if r["is_regular"] else "", r["group_name"] or ""]
            for r in usage.open_stays()]
    return _csv_response(rows, _OPEN_HEADER,
                         "parking_open_stays_%s" % usage.today_kst().isoformat())


@admin_bp.route("/overuse/export")
@admin_required
def export_overuse():
    report = usage.scan_overuse(month=request.args.get("month"))
    rows = [[i, r["car_number"], r["days"], r["over"], r["stay_text"],
             r["first_day"], r["last_day"], _fmt(r["last_event"]), r["households"],
             "예" if r["registered"] else "아니오", "예" if r["open_in"] else ""]
            for i, r in enumerate(report["rows"], 1)]
    return _csv_response(rows, _OVERUSE_HEADER, "parking_overuse_%s" % report["period"])


@admin_bp.route("/popups")
@admin_required
def popups():
    return render_template("admin/popups.html", popups=models.popups_all())


@admin_bp.route("/popups/new", methods=["POST"])
@admin_required
def popup_new():
    title = (request.form.get("title") or "").strip()
    content = (request.form.get("content") or "").strip()
    start_date = (request.form.get("start_date") or "").strip() or None
    end_date = (request.form.get("end_date") or "").strip() or None
    if not title or not content:
        flash("제목과 내용을 입력하세요.", "danger")
        return redirect(url_for("admin.popups"))
    models.popups_create({"title": title, "content": content,
                          "start_date": start_date, "end_date": end_date, "is_active": True})
    flash("팝업을 등록했습니다.", "success")
    return redirect(url_for("admin.popups"))


@admin_bp.route("/popups/<int:pid>/toggle", methods=["POST"])
@admin_required
def popup_toggle(pid):
    p = models.popups_get(pid)
    if not p:
        abort(404)
    models.popups_update(pid, {"is_active": not p.is_active})
    return redirect(url_for("admin.popups"))


@admin_bp.route("/popups/<int:pid>/delete", methods=["POST"])
@admin_required
def popup_delete(pid):
    models.popups_delete(pid)
    flash("팝업을 삭제했습니다.", "info")
    return redirect(url_for("admin.popups"))


# ---------------------------------------------------------- 경비실 차량조회
# 조회 화면 자체는 lookup 블루프린트(공용 PIN)에 있다. 여기서는 배포용 QR 과
# 조회기록만 다룬다 — 공용 PIN 이라 조회자가 계정으로 남지 않으므로, 기록을
# 근무자 배치표와 대조하는 것이 유일한 추적 수단이다.

@admin_bp.route("/lookups")
@admin_required
def lookups():
    date_from = _parse_date(request.args.get("from"))
    date_to = _parse_date(request.args.get("to"))
    # 화면에 찍히는 시각은 KST 다. 필터도 KST 하루로 잡아 준다.
    since = (date_from - timedelta(hours=9)) if date_from else None
    until = (date_to + timedelta(days=1, hours=-9)) if date_to else None
    return render_template("admin/lookups.html",
                           logs=models.lookup_logs_recent(since, until),
                           date_from=request.args.get("from", ""),
                           date_to=request.args.get("to", ""))


@admin_bp.route("/lookup-qr")
@admin_required
def lookup_qr():
    """경비실에 붙일 QR. PIN 이 박혀 있어 찍으면 곧바로 조회 화면이 열린다.

    QR 은 서버에서 SVG 로 그려 본문에 심는다. 이 화면의 결과물은 경비실 벽에
    붙는 종이라, 외부 CDN 스크립트에 매달아 두면 그 CDN 이 흔들리는 날 인쇄가
    빈 칸으로 나온다. SVG 라 확대·인쇄에도 깨지지 않는다.
    """
    from .lookup import read_pin_store, _pin as current_pin
    pin = current_pin()
    secret = (current_app.config.get("LOOKUP_QR_SECRET") or "").strip()
    base_url = url_for("lookup.index", _external=True)
    # QR 에는 PIN 이 아니라 별개의 무작위 토큰을 싣는다. PIN 을 주소에 실으면
    # 주소창을 보는 사람에게 그대로 읽히고, 쿼리스트링이면 액세스 로그에도 남는다.
    # 프래그먼트(#)는 서버로 전송되지 않아 로그·Referer 어디에도 남지 않는다.
    qr_url = base_url + ("#k=" + quote(secret, safe="") if secret else "")

    qr_svg = None
    if secret:
        try:
            import segno
            qr_svg = Markup(segno.make(qr_url, error="m").svg_inline(
                scale=6, border=2, dark="#0f172a"))
        except Exception:
            current_app.logger.exception("경비실 QR 생성 실패")

    return render_template("admin/lookup_qr.html", pin=pin, pin_set=bool(pin),
                           qr_set=bool(secret), base_url=base_url,
                           qr_url=qr_url, qr_svg=qr_svg,
                           pin_meta=read_pin_store())


# PIN 은 경비원이 외워서 치는 값이라 숫자로 제한한다. 4자리는 너무 짧아 6자리
# 이상을 권하지만, 기존 운영 번호를 그대로 옮겨 쓸 수 있게 4자리부터 허용한다.
PIN_RE = re.compile(r"^\d{4,12}$")


@admin_bp.route("/lookup-pin", methods=["POST"])
@admin_required
def lookup_pin_change():
    """경비실 PIN 교체. QR 토큰은 건드리지 않는다 — 둘은 별개의 비밀이다."""
    from .lookup import write_pin_store, client_ip
    new_pin = (request.form.get("pin") or "").strip()
    confirm = (request.form.get("pin_confirm") or "").strip()

    if not PIN_RE.match(new_pin):
        flash("PIN 은 숫자 4~12자리로 입력하세요.", "danger")
    elif new_pin != confirm:
        flash("확인용 PIN 이 일치하지 않습니다.", "danger")
    else:
        try:
            write_pin_store(new_pin, getattr(current_user, "name", ""))
        except OSError:
            current_app.logger.exception("PIN 저장 실패")
            flash("PIN 을 저장하지 못했습니다. 잠시 후 다시 시도해 주세요.", "danger")
        else:
            # 누가 언제 바꿨는지 조회기록에 남긴다. 값 자체는 남기지 않는다.
            models.lookup_log_add("pin_change", ip=client_ip())
            flash("PIN 을 변경했습니다. 경비실에 새 번호를 알려 주세요. "
                  "이미 열려 있는 경비실 화면은 그대로 쓸 수 있습니다.", "success")
    return redirect(url_for("admin.lookup_qr"))
